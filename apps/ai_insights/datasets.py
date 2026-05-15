import csv
import json
from io import StringIO
from pathlib import Path

from rest_framework import serializers


MAX_DATASET_ROWS = 500


def _normalise_row(row):
    return {
        key: value
        for key, value in row.items()
        if key is not None and value not in (None, '')
    }


def _entries_from_json_file(uploaded_file):
    try:
        payload = json.loads(uploaded_file.read().decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise serializers.ValidationError({'file': 'Uploaded JSON is invalid.'}) from exc
    return _entries_from_payload(payload)


def _entries_from_csv_file(uploaded_file):
    try:
        text = uploaded_file.read().decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise serializers.ValidationError({'file': 'Uploaded CSV must be UTF-8 encoded.'}) from exc
    reader = csv.DictReader(StringIO(text))
    return [_normalise_row(row) for row in reader]


def _entries_from_xlsx_file(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise serializers.ValidationError({
            'file': 'XLSX uploads require openpyxl to be installed.'
        }) from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    normalised_headers = [str(header).strip() if header is not None else None for header in headers]
    entries = []
    for row in rows:
        entries.append(_normalise_row(dict(zip(normalised_headers, row))))
    return entries


def _entries_from_payload(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        entries = payload.get('entries') or payload.get('data') or payload.get('rows')
        if isinstance(entries, list):
            return entries
    raise serializers.ValidationError({
        'entries': 'Provide a JSON array or an object containing an entries array.'
    })


def load_dataset_entries(request):
    uploaded_file = request.FILES.get('file')
    if uploaded_file:
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix == '.json':
            entries = _entries_from_json_file(uploaded_file)
        elif suffix == '.csv':
            entries = _entries_from_csv_file(uploaded_file)
        elif suffix in ('.xlsx', '.xlsm'):
            entries = _entries_from_xlsx_file(uploaded_file)
        else:
            raise serializers.ValidationError({
                'file': 'Unsupported file type. Upload .json, .csv, .xlsx, or .xlsm.'
            })
    else:
        entries = _entries_from_payload(request.data)

    if not entries:
        raise serializers.ValidationError({'entries': 'Dataset must contain at least one row.'})
    if len(entries) > MAX_DATASET_ROWS:
        raise serializers.ValidationError({
            'entries': f'Dataset contains {len(entries)} rows. Maximum allowed is {MAX_DATASET_ROWS}.'
        })
    if not all(isinstance(entry, dict) for entry in entries):
        raise serializers.ValidationError({'entries': 'Every dataset row must be an object.'})

    return entries
